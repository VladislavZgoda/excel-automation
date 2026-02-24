import re
from pathlib import Path
from datetime import date
from typing import cast
from sys import argv, exit

import pandas as pd
import numpy as np
from natsort import index_natsorted
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font

if len(argv) < 2:
    print(
        "Error: Missing argument! Example: 'uv run p2_readings/main.py private | legal'."
    )
    exit(1)
elif argv[1] not in ["private", "legal"]:
    print("Error: The script argument is invalid! It must be 'private' or 'legal'.")
    exit(1)

is_private = True if argv[1] == "private" else False
script_dir = Path(__file__).resolve().parent
p2_readings_path = script_dir / "input_files" / "p2_readings.xlsx"
output_path = script_dir / "output_files"


def extract_date(header: str) -> str:
    pattern = r"^(?:Показание на начало периода|Показание на конец периода) (.*)$"
    m = re.match(pattern, header)

    if m:
        return m.group(1)
    return header


df = pd.read_excel(p2_readings_path, header=[0, 1], skiprows=4)

level0_headers = df.columns.get_level_values(0)
level1_headers = df.columns.get_level_values(1)
new_level0_headers = [extract_date(val) for val in level0_headers]

df.columns = pd.MultiIndex.from_arrays([new_level0_headers, level1_headers])


def copy_by_date(row: pd.Series) -> pd.Series:
    current_date: str = row[("Дата последних показаний", "Unnamed: 8_level_1")]
    columns: list[str] = ["Т1", "Т2", "Т3", "Общая"]

    for col in columns:
        row[col] = row[(current_date, col)]

    return row


df: pd.DataFrame = df.apply(copy_by_date, axis="columns")

df = df[
    [
        "№пп",
        "Серийный номер",
        "Дата последних показаний",
        "Т1",
        "Т2",
        "Т3",
        "Общая",
        "Точка учёта",
        "Абонент",
        "Тип",
        "Наименование балансовой группы",
    ]
]

df.columns = df.columns.droplevel(1)

df.columns = [
    "№ п/п",
    "Номер_ПУ",
    "Дата",
    "Т1",
    "Т2",
    "Т3",
    "Т сумм",
    "Адрес",
    "ФИО абонента",
    "Тип ПУ",
    "ТП",
]

df.insert(1, "Л/С", None)
df["Л/С"] = df["ФИО абонента"].str.extract(r"(\d{12})")[0]

# [А-ЯЁ][а-яё]+ - первая часть фамилии.
# (?:-[А-ЯЁ][а-яё]+)? - необязательный дефис со второй частью фамилии, когда из двух частей.
# \s+ - один или несколько пробелов между фамилией и инициалами.
# [А-ЯЁ]\. - первый инициал с точкой.
# (?:\s?[А-ЯЁ]\.)? - возможный пробел и второй инициал с точкой.
pattern = r"^([А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?\s+[А-ЯЁ]\.(?:\s?[А-ЯЁ]\.)?)"
df["ФИО абонента"] = df["ФИО абонента"].str.extract(pattern)

# Убрать модель ПУ и S/N.
df["Адрес"] = df["Адрес"].str.split("\\").apply(lambda x: ", ".join(x[:-1]))

# Добавить 0 к началу серийного номера, если он из 7 цифр.
df["Номер_ПУ"] = df["Номер_ПУ"].apply(str).str.zfill(8)

df["Т1"] = df["Т1"].round(2).replace("н/д", None)
df["Т2"] = df["Т2"].round(2).replace("н/д", None)
df["Т3"] = df["Т3"].round(2).replace("н/д", None)
df["Т сумм"] = df["Т сумм"].round(2).replace("н/д", None)

askue_date = date.today().strftime("%d.%m.%Y")
df.insert(10, "Дата_АСКУЭ", askue_date)
df.insert(12, "Способ снятия показаний", "УСПД")

# Естественный порядок сортировки
df = df.iloc[index_natsorted(df["ТП"])]

# Добавить нумерацию для строк.
df["№ п/п"] = np.arange(1, len(df) + 1)

alignment_center = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)

alignment_left = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
)

alignment_right = Alignment(
    horizontal="right",
    vertical="center",
    wrap_text=True,
)

border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin"),
)

font = Font(name="Times New Roman", size=10)
font_bold = Font(name="Times New Roman", size=10, bold="True")
font_title = Font(name="Times New Roman", size=12, bold="True")


def write_df_to_wb(df: pd.DataFrame) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "БЫТ" if is_private else "ЮР"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    return wb, ws


def style_ws(ws: Worksheet, title_range: str) -> None:
    for col in ws.iter_cols():
        for cell in col:
            cell.border = border
            cell.number_format = "@"

            column_letter = cell.column_letter

            if column_letter in ["E", "F", "G", "H"]:
                cell.alignment = alignment_right
            elif column_letter in ["I", "J"]:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_center

            cell_row_number = cast(int, cell.row)

            if cell_row_number == 1:
                cell.font = font_bold
            else:
                cell.font = font

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 25

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = None

    ws.insert_rows(1)
    ws.merge_cells(title_range)
    ws[
        "A1"
    ].value = "Ведомость дистанционного снятия показаний посредствам АСКУЭ и ридера"
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center


wb, ws = write_df_to_wb(df)
style_ws(ws=ws, title_range="A1:N1")

wb.save(
    output_path / f"Приложение №9 {f'Быт {askue_date}' if is_private else 'Юр'}.xlsx"
)

if is_private:
    df.drop(
        ["Дата_АСКУЭ", "Тип ПУ", "Способ снятия показаний", "ТП"], axis=1, inplace=True
    )

    df["Ведомость_КС"] = ""
    df["Контролер"] = "Згода В.Г."

    wb, ws = write_df_to_wb(df)
    style_ws(ws=ws, title_range="A1:L1")
    wb.save(output_path / f"АСКУЭ Быт {askue_date}.xlsx")
