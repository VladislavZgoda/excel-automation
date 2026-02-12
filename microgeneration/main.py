from pathlib import Path
from sys import argv, exit
from datetime import datetime
from typing import cast, TypedDict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

script_dir = Path(__file__).resolve().parent

if len(argv) < 2:
    print(
        "Error: Missing argument! Example: 'uv run microgeneration/main.py private | legal'."
    )
    exit(1)
elif argv[1] not in ["private", "legal"]:
    print("Error: The script argument is invalid! It must be 'private' or 'legal'.")
    exit(1)


# Прочитать выгрузку из Sims.
try:
    wb_matritca_readings = load_workbook(
        script_dir / "input_files" / "matritca_readings.xlsx"
    )
except FileNotFoundError:
    print("FileNotFoundError: The file 'matritca_readings.xlsx' not found.")
    exit(1)

ws_matritca_readings = cast(Worksheet, wb_matritca_readings.active)

# Определить шаблон для чтения.
try:
    if argv[1] == "private":
        wb_template = load_workbook(script_dir / "templates" / "private.xlsx")
    elif argv[1] == "legal":
        wb_template = load_workbook(script_dir / "templates" / "legal.xlsx")
except FileNotFoundError:
    print("FileNotFoundError: The file 'private.xlsx or legal.xlsx' not found.")
    exit(1)

ws_template = cast(Worksheet, wb_template.active)

MeterData = TypedDict(
    "MeterData",
    {
        "T1_import": int | float | None,
        "T2_import": int | float | None,
        "T_import": int | float | None,
        "T1_export": int | float | None,
        "T2_export": int | float | None,
        "T_export": int | float | None,
        "date": datetime,
    },
)

meters_dict: dict[str, MeterData] = {}
current_date = datetime.today().strftime("%d.%m.%Y")

# Создать словарь с ключом в виде s/n ПУ из шаблона.
for row in range(3, ws_template.max_row + 1):
    str_row = str(row)
    meter: str = ws_template["C" + str_row].value
    meters_dict[meter] = {}

# Заполнить meter_dict данными.
for row in range(3, ws_matritca_readings.max_row + 1):
    str_row = str(row)
    meter = str(ws_matritca_readings["C" + str_row].value)

    if len(meter) == 7:
        meter = "0" + meter

    if meter not in meters_dict:
        continue

    meters_dict[meter]["T1_import"] = ws_matritca_readings["E" + str_row].value
    meters_dict[meter]["T2_import"] = ws_matritca_readings["F" + str_row].value
    meters_dict[meter]["T_import"] = ws_matritca_readings["H" + str_row].value
    meters_dict[meter]["T1_export"] = ws_matritca_readings["I" + str_row].value
    meters_dict[meter]["T2_export"] = ws_matritca_readings["J" + str_row].value
    meters_dict[meter]["T_export"] = ws_matritca_readings["L" + str_row].value
    meters_dict[meter]["date"] = ws_matritca_readings["D" + str_row].value


def try_round_value(val: int | float | None) -> float | None:
    if val:
        return round(val, 2)


# Записать данные в шаблон.
for row in range(3, ws_template.max_row + 1):
    str_row = str(row)
    meter = ws_template["C" + str_row].value

    if not meters_dict[meter]:
        continue

    reading_date = meters_dict[meter]["date"].strftime("%d.%m.%Y")
    ws_template["D" + str_row].value = reading_date
    ws_template["E" + str_row].value = try_round_value(meters_dict[meter]["T1_import"])
    ws_template["F" + str_row].value = try_round_value(meters_dict[meter]["T2_import"])
    ws_template["H" + str_row].value = try_round_value(meters_dict[meter]["T_import"])
    ws_template["I" + str_row].value = try_round_value(meters_dict[meter]["T1_export"])
    ws_template["I" + str_row].value = try_round_value(meters_dict[meter]["T1_export"])
    ws_template["J" + str_row].value = try_round_value(meters_dict[meter]["T2_export"])
    ws_template["L" + str_row].value = try_round_value(meters_dict[meter]["T_export"])
    ws_template["O" + str_row].value = current_date


wb_template.save(
    script_dir
    / "output_files"
    / f"Микрогенерация {'Быт' if argv[1] == 'private' else 'Юр'}.xlsx"
)
