import re
from sys import exit
from pathlib import Path
from typing import TypedDict
from time import strftime, localtime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

script_dir = Path(__file__).resolve().parent

# meter_readings.xlsx - Отчет "Новые показания" из Пирамида 2.
meter_readings_path = script_dir / "input_files" / "meter_readings.xlsx"

try:
    wb_meter_readings = load_workbook(meter_readings_path)
    ws_meter_readings = wb_meter_readings["Данные"]
except FileNotFoundError:
    print("FileNotFoundError: The file 'meter_readings.xlsx' not found.")
    exit(1)
except KeyError:
    print("KeyError: The sheet 'Данные' was not found.")
    exit(1)

# current_meter_reading.xlsx - Выгрузка показаний из Пирамида 2 с А+ текущие.
current_meter_readings_path = script_dir / "input_files" / "current_meter_readings.xlsx"

try:
    wb_current_meter_readings = load_workbook(current_meter_readings_path)
    ws_current_meter_readings = wb_current_meter_readings["Sheet"]
except FileNotFoundError:
    print("FileNotFoundError: The file 'current_meter_readings.xlsx' not found.")
    exit(1)
except KeyError:
    print("KeyError: The sheet 'Sheet' was not found.")
    exit(1)

# Из matritca_readings.xlsx интересуют только ВИП потребители.
matritca_readings_path = script_dir / "input_files" / "matritca_readings.xlsx"

try:
    # Пропустить первую строку, чтобы не использовать ее в качестве headers.
    df = pd.read_excel(matritca_readings_path, skiprows=[0])
except FileNotFoundError:
    print("FileNotFoundError: The file 'matritca_readings.xlsx' not found.")
    exit(1)

# Удалить последнюю строку.
df = df.iloc[:-1]

df["Код потребителя"] = df["Код потребителя"].str.extract(r"(\d{12})")
df = df[df["Код потребителя"].notna()]
df = df[~df["Код потребителя"].str[:6].isin(["230700", "230710"])]

# Убрать повторяющиеся строки с ПУ у которых выгрузились Т1-Т2 и Т общ отдельно.
df = df[df["Активная энергия, импорт"].notna()]

df["Дата"] = df["Дата"].dt.strftime("%d.%m.%Y")

df = df.drop_duplicates(subset=["Серийный №"], keep="last")
df["Серийный №"] = df["Серийный №"].astype("int").astype("str")

# Добавить 0 к началу серийного номера, если он из 7 цифр.
df["Серийный №"] = df["Серийный №"].str.zfill(8)
df = df.set_index("Серийный №")[["Дата", "Активная энергия, импорт"]]

MeterData = TypedDict(
    "MeterData",
    {
        "readings": int | float,
        "date": str,
    },
)

meter_readings: dict[str, MeterData] = {}

for serial_number, values in df.to_dict(orient="index").items():
    meter_readings[str(serial_number)] = {
        "date": str(values["Дата"]),
        "readings": float(values["Активная энергия, импорт"]),
    }

current_date = strftime("%d.%m.%Y", localtime())
meter_readings_date = ws_meter_readings["K6"].value

alignment_date = Alignment(horizontal="center", vertical="center")
alignment_value = Alignment(horizontal="right", vertical="center")

for row in range(7, ws_meter_readings.max_row + 1):
    str_row_number = str(row)
    readings = ws_meter_readings["K" + str_row_number].value

    if not isinstance(readings, (int, float)):
        continue

    serial_number = ws_meter_readings["E" + str_row_number].value
    meter_readings[serial_number] = {
        "readings": readings,
        "date": meter_readings_date,
    }

serial_number_regex = re.compile(r"\d{6,8}")

for row in range(3, ws_current_meter_readings.max_row + 1):
    str_row_number = str(row)
    serial_number_cell = ws_current_meter_readings["A" + str_row_number].value
    serial_number = serial_number_regex.search(serial_number_cell)

    if serial_number is None:
        continue

    readings = ws_current_meter_readings["C" + str_row_number].value

    if not isinstance(readings, (int, float)):
        continue

    meter_readings[serial_number.group()] = {
        "readings": readings,
        "date": current_date,
    }

templates_dir = Path(script_dir / "templates")

for file in templates_dir.iterdir():
    wb = load_workbook(file)
    ws = wb.active

    if ws is None:
        continue

    for row in range(3, ws.max_row + 1):
        str_row_number = str(row)
        serial_number = str(ws["C" + str_row_number].value)

        if serial_number not in meter_readings:
            continue

        meter_data = meter_readings[serial_number]
        ws["H" + str_row_number].value = round(meter_data["readings"], 2)
        ws["H" + str_row_number].alignment = alignment_value
        ws["D" + str_row_number].value = meter_data["date"]
        ws["D" + str_row_number].alignment = alignment_date
        ws["K" + str_row_number].value = current_date
        ws["K" + str_row_number].alignment = alignment_date

    wb.save(script_dir / "output_files" / file.name)
