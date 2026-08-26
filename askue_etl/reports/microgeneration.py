from datetime import datetime
from pathlib import Path
from typing import cast
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..readings.microgeneration import MeterReadings


def write_readings_report(template_path: Path, readings: MeterReadings) -> Workbook:
    wb = load_workbook(template_path)
    ws = cast(Worksheet, wb.active)

    askue_date = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")

    for row in range(3, ws.max_row + 1):
        str_row = str(row)
        meter = ws["C" + str_row].value
        meter_readings = readings.get(meter)

        if meter_readings is None:
            continue

        ws["D" + str_row].value = meter_readings.date.strftime("%d.%m.%Y")
        ws["E" + str_row].value = _try_round_value(meter_readings.T1_import)
        ws["F" + str_row].value = _try_round_value(meter_readings.T2_import)
        ws["H" + str_row].value = _try_round_value(meter_readings.T_import)
        ws["I" + str_row].value = _try_round_value(meter_readings.T1_export)
        ws["J" + str_row].value = _try_round_value(meter_readings.T2_export)
        ws["L" + str_row].value = _try_round_value(meter_readings.T_export)
        ws["O" + str_row].value = askue_date

    return wb


def _try_round_value(val: float | None) -> float | None:
    if val is None:
        return None

    return round(val, 2)
