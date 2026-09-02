from datetime import datetime
from pathlib import Path
from typing import cast
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..readings.missing_readings import MeterReadings


def write_readings_report(report_path: Path, meter_readings: MeterReadings) -> Workbook:
    wb = load_workbook(report_path)
    ws = cast(Worksheet, wb.active)

    askue_date = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")

    alignment_date = Alignment(horizontal="center", vertical="center")
    alignment_value = Alignment(horizontal="right", vertical="center")

    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    for meter_data in meter_readings:
        serial_number = meter_data.serial_number
        row_number = meter_data.row_number
        readings = meter_data.readings
        readings_date = meter_data.readings_date

        cell_serial_number = str(ws["C" + row_number].value)

        if serial_number != cell_serial_number:
            raise ValueError(
                f"Несовпадение счётчика в строке {row_number}: "
                f"ожидался {serial_number!r}, найден {cell_serial_number!r} в {report_path.name}."
            )

        cell_d = ws["D" + row_number]
        cell_h = ws["H" + row_number]
        cell_k = ws["K" + row_number]

        cell_h.value = round(readings, 2)
        cell_h.alignment = alignment_value
        cell_h.fill = yellow_fill
        cell_d.value = readings_date
        cell_d.alignment = alignment_date
        cell_k.value = askue_date
        cell_k.alignment = alignment_date

    return wb
