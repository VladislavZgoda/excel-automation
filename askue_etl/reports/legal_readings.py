from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from askue_etl.readings.legal_readings import MeterReadings


def write_readings_reports(
    meter_readings: MeterReadings, template_folder_path: Path, reports_folder_path: Path
) -> None:
    current_date = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")

    alignment_date = Alignment(horizontal="center", vertical="center")
    alignment_value = Alignment(horizontal="right", vertical="center")

    for file in template_folder_path.iterdir():
        wb = load_workbook(file)
        ws = wb.active

        if ws is None:
            continue

        for row in range(3, ws.max_row + 1):
            str_row_number = str(row)
            serial_number = str(ws["C" + str_row_number].value)

            if serial_number not in meter_readings:
                continue

            meter_data = meter_readings[serial_number]
            ws["H" + str_row_number].value = round(meter_data.readings, 2)
            ws["H" + str_row_number].alignment = alignment_value
            ws["D" + str_row_number].value = meter_data.date
            ws["D" + str_row_number].alignment = alignment_date
            ws["K" + str_row_number].value = current_date
            ws["K" + str_row_number].alignment = alignment_date

        wb.save(reports_folder_path / file.name)
