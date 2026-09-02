from dataclasses import dataclass
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# key - серийный номер счётчика
# value - номер строки в приложении №9
type MissingReadingsMeters = dict[str, str]


@dataclass(frozen=True)
class MeterData:
    serial_number: str
    readings: int | float
    row_number: str


type MeterReadings = list[MeterData]


def prepare_readings(readings_path: Path, report_path: Path) -> MeterReadings:
    meters = _get_meters_without_readings(report_path)
    meter_readings = _get_readings(readings_path, meters)

    return meter_readings


def _get_meters_without_readings(report_path: Path) -> MissingReadingsMeters:
    wb = load_workbook(report_path, read_only=True)
    ws = cast(Worksheet, wb.active)

    meters: MissingReadingsMeters = {}

    for row in ws.iter_rows(min_row=3, min_col=3, max_col=8):
        readings = row[5].value  # H (индекс 5: C=0, D=1, E=2, F=3, G=4, H=5)

        if readings is None:
            serial_number = str(row[0].value)
            meters[serial_number] = str(row[0].row)

    wb.close()
    return meters


def _get_readings(readings_path: Path, meters: MissingReadingsMeters) -> MeterReadings:
    wb = load_workbook(readings_path, read_only=True)
    ws = cast(Worksheet, wb.active)

    meter_readings: MeterReadings = []

    for row in ws.iter_rows(min_row=7, min_col=5, max_col=11):
        serial_number = str(row[0].value).zfill(8)  # E

        if serial_number not in meters:
            continue

        readings = row[6].value  # K

        if not isinstance(readings, (int, float)):
            continue

        meter_readings.append(
            MeterData(
                serial_number=serial_number,
                readings=readings,
                row_number=meters[serial_number],
            )
        )

    wb.close()
    return meter_readings
