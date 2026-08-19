import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import polars as pl
from openpyxl import load_workbook


@dataclass(frozen=True)
class MeterData:
    date: str
    readings: int | float


type MeterReadings = dict[str, MeterData]


def prepare_readings(
    sims_readings_path: Path, p2_readings_path: Path, p2_current_readings: Path
) -> MeterReadings:
    meter_readings: MeterReadings = {}

    meter_readings |= _get_matritca_vip_readings(sims_readings_path)
    meter_readings |= _get_p2_readings(p2_readings_path)
    meter_readings |= _get_p2_current_readings(p2_current_readings)

    return meter_readings


def _get_matritca_vip_readings(readings_path: Path) -> MeterReadings:
    df = (
        pl.read_excel(readings_path, read_options={"header_row": 1})
        .head(-1)
        .with_columns(pl.col("Код потребителя").str.extract(r"(\d{12})"))
        .filter(
            pl.col("Код потребителя").is_not_null(),
            ~pl.col("Код потребителя").str.slice(0, 5).is_in(["23070", "23071"]),
            pl.col("Активная энергия, импорт").is_not_null(),
        )
        .with_columns(
            pl.col("Серийный №").str.zfill(8),
            pl.col("Дата").dt.to_string("%d.%m.%Y"),
        )
        .select(["Серийный №", "Дата", "Активная энергия, импорт"])
    )

    return {
        record["Серийный №"]: MeterData(
            date=record["Дата"],
            readings=float(record["Активная энергия, импорт"]),
        )
        for record in df.iter_rows(named=True)
    }


def _get_p2_readings(readings_path: Path) -> MeterReadings:
    wb = load_workbook(readings_path)
    ws = wb["Данные"]

    meter_readings: MeterReadings = {}
    readings_date = str(ws["K6"].value)

    for row in range(7, ws.max_row + 1):
        str_row_number = str(row)
        readings = ws["K" + str_row_number].value

        if not isinstance(readings, (int, float)):
            continue

        serial_number = str(ws["E" + str_row_number].value)
        meter_readings[serial_number] = MeterData(
            date=readings_date,
            readings=readings,
        )

    return meter_readings


def _get_p2_current_readings(readings_path: Path) -> MeterReadings:
    wb = load_workbook(readings_path)
    ws = wb["Sheet"]

    meter_readings: MeterReadings = {}
    current_date = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")
    serial_number_regex = re.compile(r"\d{6,8}")

    for row in range(3, ws.max_row + 1):
        str_row_number = str(row)
        serial_number_cell = ws["A" + str_row_number].value
        serial_number = serial_number_regex.search(serial_number_cell)

        if serial_number is None:
            continue

        readings = ws["C" + str_row_number].value

        if not isinstance(readings, (int, float)):
            continue

        meter_readings[serial_number.group()] = MeterData(
            date=current_date,
            readings=readings,
        )

    return meter_readings
