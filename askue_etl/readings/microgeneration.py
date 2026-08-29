from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class MeterData:
    date: datetime
    T1_import: float | None
    T2_import: float | None
    T_import: float | None
    T1_export: float | None
    T2_export: float | None
    T_export: float | None


type MeterReadings = dict[str, MeterData]
type Meters = set[str]


def prepare_readings(readings_path: Path, template_path: Path) -> MeterReadings:
    meters = _get_meters_from_template(template_path)
    meter_readings = _get_readings(readings_path, meters)
    return meter_readings


def _get_meters_from_template(template_path: Path) -> Meters:
    wb = load_workbook(template_path)
    ws = cast(Worksheet, wb.active)

    meters: Meters = set()

    for row in range(3, ws.max_row + 1):
        str_row = str(row)
        meter = str(ws["C" + str_row].value)
        meters.add(meter)

    return meters


def _get_readings(readings_path: Path, meters: Meters) -> MeterReadings:
    wb = load_workbook(readings_path, read_only=True)
    ws = cast(Worksheet, wb.active)

    meter_readings: MeterReadings = {}

    for row in ws.iter_rows(min_row=3, min_col=3, max_col=12):
        row_number = cast(int, row[0].row)
        meter = str(row[0].value).zfill(8)  # C

        if meter not in meters:
            continue

        meter_readings[meter] = MeterData(
            date=_to_datetime(row[1].value, row_number, "D"),
            T1_import=_to_float(row[2].value, row_number, "E"),
            T2_import=_to_float(row[3].value, row_number, "F"),
            T_import=_to_float(row[5].value, row_number, "H"),
            T1_export=_to_float(row[6].value, row_number, "I"),
            T2_export=_to_float(row[7].value, row_number, "J"),
            T_export=_to_float(row[9].value, row_number, "L"),
        )

    wb.close()
    return meter_readings


def _to_float(value: object, row: int, column: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    raise TypeError(
        f"Ожидалось числовое значение в ячейке {column}{row}, "
        f"получено {value!r} ({type(value).__name__})"
    )


def _to_datetime(value: object, row: int, column: str) -> datetime:
    if isinstance(value, datetime):
        return value

    raise TypeError(
        f"Ожидалась дата (datetime) в ячейке {column}{row}, "
        f"получено {value!r} ({type(value).__name__})"
    )
