from sys import exit
from pathlib import Path
from time import strftime, localtime
from typing import TypedDict, TypeAlias

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

script_dir = Path(__file__).resolve().parent

# read_data.xlsx - Отчет "Новые показания" из Пирамида 2.
read_data_path = script_dir / "input_files" / "read_data.xlsx"
# write_data.xlsx - Приложение №9.
write_data_path = script_dir / "input_files" / "write_data.xlsx"

try:
    wb_read_data = load_workbook(read_data_path)
    ws_read_data = wb_read_data["Данные"]
except FileNotFoundError:
    print("FileNotFoundError: The file 'read_data.xlsx' not found.")
    exit(1)
except KeyError:
    print("KeyError: The sheet 'Данные' was not found.")
    exit(1)

try:
    wb_write_data = load_workbook(write_data_path)
    ws_write_data = wb_write_data["ЮР"]
except FileNotFoundError:
    print("FileNotFoundError: The file 'write_data.xlsx' not found.")
    exit(1)
except KeyError:
    print("KeyError: The sheet 'ЮР' was not found.")
    exit(1)

current_date = strftime("%d.%m.%Y", localtime())
meter_reading_date = ws_read_data["K6"].value

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                          fill_type="solid")

MeterData = TypedDict(
    "MeterData",
    {
        "serial_number": str,
        "readings": int | float,
        "row_number": str,
    },
)

RowNumber = TypedDict("RowNumber", {"row_number": str})

MeterReadings: TypeAlias = list[MeterData]

# key - Серийный номер ПУ.
meters_without_readings: dict[str, RowNumber] = {}
meters_readings: MeterReadings = []

for row in range(3, ws_write_data.max_row + 1):
    str_row_number = str(row)
    readings = ws_write_data["H" + str_row_number].value

    if readings is None:
        serial_number = ws_write_data["C" + str_row_number].value
        meters_without_readings[serial_number] = {"row_number": str_row_number}

for row in range(7, ws_read_data.max_row + 1):
    str_row_number = str(row)
    serial_number = ws_read_data["E" + str_row_number].value

    if serial_number not in meters_without_readings:
        continue

    readings = ws_read_data["K" + str_row_number].value

    if not isinstance(readings, (int, float)):
        continue

    meter_data: MeterData = {
        "serial_number": serial_number,
        "readings": readings,
        "row_number": meters_without_readings[serial_number]["row_number"],
    }
    meters_readings.append(meter_data)

for meter_data in meters_readings:
    serial_number = meter_data["serial_number"]
    row_number = meter_data["row_number"]
    readings = meter_data["readings"]
    cell_serial_number = ws_write_data["C" + row_number].value

    if meter_data["serial_number"] != cell_serial_number:
        print(f"{serial_number} not fond in the 'write_data.xlsx'.")
        exit(1)

    ws_write_data["H" + row_number].value = round(readings, 2)
    ws_write_data["D" + row_number].value = meter_reading_date
    ws_write_data["K" + row_number].value = current_date
    ws_write_data["H" + row_number].fill = yellow_fill

wb_write_data.save(script_dir / "output_files" / "Приложение №9 ЮР.xlsx")
