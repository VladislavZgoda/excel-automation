from typing import cast
from pathlib import Path
from sys import argv, exit
from datetime import date

import pandas as pd
import numpy as np
from natsort import index_natsorted
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font

if len(argv) < 2:
    print(
        "Error: Missing argument! Example: 'uv run process_matritca_readings/main.py private | legal'."
    )
    exit(1)
elif argv[1] not in ["private", "legal"]:
    print("Error: The script argument is invalid! It must be 'private' or 'legal'.")
    exit(1)

script_dir = Path(__file__).resolve().parent

# matritca_readings.xlsx - Выгрузка показаний из Sims Client.
matritca_readings_path = script_dir / "input_files" / "matritca_readings.xlsx"
output_path = script_dir / "output_files"

is_private = True if argv[1] == "private" else False
consumer_number_filter = "230700" if is_private else "230710"

alignment_center = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)

alignment_left = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
)

alignment_right = Alignment(
    horizontal="right",
    vertical="center",
    wrap_text=True,
)

border = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin"),
)

font = Font(name="Times New Roman", size=10)
font_bold = Font(name="Times New Roman", size=10, bold="True")
font_title = Font(name="Times New Roman", size=12, bold="True")

# Пропустить первую строку, чтобы не использовать ее в качестве headers.
df = pd.read_excel(matritca_readings_path, skiprows=[0])
df.columns = [
    "№ п/п",
    "Л/С",
    "Номер_ПУ",
    "Дата",
    "Т1",
    "Т2",
    "Т3",
    "Т сумм",
    "Адрес",
    "ФИО абонента",
    "Тип ПУ",
]

# Удалить последнюю строку.
df = df.iloc[:-1]

df["Л/С"] = df["Л/С"].str.extract(r"(\d{12})")
df = df[df["Л/С"].notna()]
df = df[df["Л/С"].str.startswith(consumer_number_filter)]

if is_private:
    df = df[df["ФИО абонента"] != "ОДПУ"]

df["Номер_ПУ"] = df["Номер_ПУ"].astype("int").astype("str")

# Добавить 0 к началу серийного номера, если он из 7 цифр.
df["Номер_ПУ"] = df["Номер_ПУ"].str.zfill(8)

df["Дата"] = df["Дата"].dt.strftime("%d.%m.%Y")

askue_date = date.today().strftime("%d.%m.%Y")
df.insert(10, "Дата_АСКУЭ", askue_date)
df["Способ снятия показаний"] = "УСПД"
df["ТП"] = df["Адрес"].str.extract(r"(ТП-\d{1,3}(П)?)", expand=False)[0]

df["Т1"] = df["Т1"].round(2)
df["Т2"] = df["Т2"].round(2)
df["Т3"] = df["Т3"].round(2)
df["Т сумм"] = df["Т сумм"].round(2)

# Естественный порядок сортировки
df = df.iloc[index_natsorted(df["ТП"])]

# Добавить нумерацию для строк.
df["№ п/п"] = np.arange(1, len(df) + 1)


def write_df_to_wb(df: pd.DataFrame) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "БЫТ" if is_private else "ЮР"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    return wb, ws


def style_ws(ws: Worksheet, title_range: str) -> None:
    for col in ws.iter_cols():
        for cell in col:
            cell.border = border
            cell.number_format = "@"

            column_letter = cell.column_letter

            if column_letter in ["E", "F", "G", "H"]:
                cell.alignment = alignment_right
            elif column_letter in ["I", "J"]:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_center

            cell_row_number = cast(int, cell.row)

            if cell_row_number == 1:
                cell.font = font_bold
            else:
                cell.font = font

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["L"].width = 18

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = None

    ws.insert_rows(1)
    ws.merge_cells(title_range)
    ws[
        "A1"
    ].value = "Ведомость дистанционного снятия показаний посредствам АСКУЭ и ридера"
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center


wb, ws = write_df_to_wb(df)
style_ws(ws=ws, title_range="A1:N1")

wb.save(
    output_path / f"Приложение №9 {f'Быт {askue_date}' if is_private else 'Юр'}.xlsx"
)

if is_private:
    df.drop(
        ["Дата_АСКУЭ", "Тип ПУ", "Способ снятия показаний", "ТП"], axis=1, inplace=True
    )

    df["Ведомость_КС"] = ""
    df["Контролер"] = "Згода В.Г."

    wb, ws = write_df_to_wb(df)
    style_ws(ws=ws, title_range="A1:L1")
    wb.save(output_path / f"АСКУЭ Быт {askue_date}.xlsx")
